import openpyxl
import pprint

# data = {id_: {
#               "info": (name_, gender_: xxx, age_: xx},
#               date_: tx_,
#                       ...
#               }
#         }
#         ...
#        }
data = {}

if __name__ == '__main__':
    wb = openpyxl.load_workbook('D:\\test.xlsx')
    for sh in wb:
        rightlowercorner = 'E' + str(sh.max_row)
        #for row in sh.iter_rows(min_row=2, max_col=3, max_row=sh.max_row, values_only=True):
        for row in sh.iter_rows(min_row=2, max_col=9, max_row=3, values_only=True):
            date_, id_, name_, gender_, age_, sx_, duration_, dx_, tx_ = row
            # Make sure the key for this id exists
            data.setdefault(id_, {"info": (name_, gender_, age_)})
            data[id_][date_] = tx_
    pprint.pprint(data)

