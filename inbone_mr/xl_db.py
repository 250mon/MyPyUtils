import openpyxl
import pt_data


class XlDb:
    def __init__(self, filename):
        self.filename = filename
        self.load_wb()

    # if a file with the filename exists, just load it
    # if not, create a new one
    def load_wb(self):
        try:
            wb = openpyxl.load_workbook(self.filename)
        except FileNotFoundError:
            wb = openpyxl.Workbook(self.filename)
        self.wb = wb

    # if a sheet with the sht_title exists, just return it
    # if not, create a new one
    def create_sht_aux(self, sht_title):
        if sht_title is None:
            return self.wb.create_sheet()
        for shtname in self.wb.sheetnames:
            if shtname == sht_title:
                return self.wb[shtname]
        return self.wb.create_sheet(sht_title)

    def create_sht(self, sht_title=None, **kwargs):
        worksht = self.create_sht_aux(sht_title)
        for k, v in kwargs.items():
            if k == 'col_headings':
                self.write_col_headings(worksht, v)
        return worksht

    def write_col_headings(self, worksht, headings):
        worksht.append(headings)

    def save_wb(self):
        self.wb.save(filename=self.filename)

    def find_row(self, worksht, pt_data):
        for row_num in range(1, worksht.max_row + 1):
            if worksht['B' + str(row_num)].value == pt_data.get_id():
                return row_num
        return None

    # write record of pt_data to the block if it is empty
    def write_record_by_id(self, worksht, pt_data, **kwargs):
        row = str(self.find_row(worksht, pt_data))
        for k, v in kwargs.items():
            if k == 'contents' and v:
                cell = worksht['G' + row]
                if not cell.value:
                    cell.value = pt_data.get_contents()

    def write_record(self, worksht, pt_data):
        worksht.append(pt_data.get_data_list())


if __name__ == '__main__':
    ptdata = pt_data.PtData('2020 6', '1000')

    headings = ['date', 'id', 'sn', 'name', 'gender', 'age', 'contents']
    ptdata.set_name("mike")
    ptdata.set_socialnumber("abd")
    ptdata.set_contents("I am a man")

    xl = XlDb('D:\\test_xldb.xlsx')
    ws1 = xl.create_sht('1', col_headings=headings)
    xl.write_record(ws1, ptdata)
    xl.save_wb()

    ptdata.set_contents("you are a boy")

    xl = XlDb('D:\\test_xldb.xlsx')
    ws = xl.create_sht('1')
    xl.write_record_by_id(ws, ptdata, contents=True)

    ws2 = xl.create_sht('2', col_headings=headings)
    xl.write_record(ws2, ptdata)
    xl.save_wb()
